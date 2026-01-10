"""
Excel Pricing & Payoff Model Generator
Creates professional Excel models for trade analysis
"""

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import LineChart, Reference, BarChart
from openpyxl.utils.dataframe import dataframe_to_rows

class ExcelPricingModel:
    """
    Generate Excel pricing models with formulas and charts
    """
    
    def __init__(self):
        self.wb = Workbook()
        
    def create_directional_trade_model(self, 
                                      trade_name="Long Copper",
                                      entry_price=8650,
                                      target_price=9200,
                                      stop_price=8400,
                                      notional=1000000):
        """
        Create directional trade pricing model
        """
        ws = self.wb.active
        ws.title = "Directional Trade"
        
        # Header
        ws['A1'] = 'DIRECTIONAL TRADE PRICING MODEL'
        ws['A1'].font = Font(size=16, bold=True, color='FFFFFF')
        ws['A1'].fill = PatternFill(start_color='1e3a8a', end_color='1e3a8a', fill_type='solid')
        ws.merge_cells('A1:F1')
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 30
        
        # Trade parameters
        ws['A3'] = 'TRADE PARAMETERS'
        ws['A3'].font = Font(bold=True, size=12)
        
        params = [
            ['Trade Name:', trade_name],
            ['Entry Price:', entry_price],
            ['Target Price:', target_price],
            ['Stop Loss:', stop_price],
            ['Notional (USD):', notional],
            ['', ''],
            ['Risk/Reward:', '=ABS((C6-C5)/(C7-C5))'],
            ['Max Profit:', '=(C6-C5)/C5'],
            ['Max Loss:', '=(C7-C5)/C5']
        ]
        
        for i, (label, value) in enumerate(params, start=4):
            ws[f'A{i}'] = label
            ws[f'A{i}'].font = Font(bold=True)
            ws[f'C{i}'] = value
            if i >= 11:
                ws[f'C{i}'].number_format = '0.00%'
        
        # Price scenarios
        ws['A15'] = 'SCENARIO ANALYSIS'
        ws['A15'].font = Font(bold=True, size=12)
        
        # Headers
        headers = ['Price', '% Change', 'P&L ($)', 'Return %', 'Status']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=16, column=col)
            cell.value = header
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='2563eb', end_color='2563eb', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')
        
        # Price scenarios
        prices = np.arange(entry_price - 500, entry_price + 800, 100)
        
        for i, price in enumerate(prices, start=17):
            ws[f'A{i}'] = price
            ws[f'B{i}'] = f'=(A{i}-$C$5)/$C$5'
            ws[f'C{i}'] = f'=(A{i}-$C$5)*($C$8/$C$5)'
            ws[f'D{i}'] = f'=C{i}/$C$8'
            ws[f'E{i}'] = (f'=IF(A{i}>=$C$6,"TARGET",IF(A{i}<=$C$7,"STOP","ACTIVE"))')
            
            # Formatting
            ws[f'B{i}'].number_format = '0.00%'
            ws[f'C{i}'].number_format = '$#,##0'
            ws[f'D{i}'].number_format = '0.00%'
        
        # Conditional formatting colors
        for i in range(17, 17 + len(prices)):
            for col in ['C', 'D']:
                cell = ws[f'{col}{i}']
                if col == 'C':
                    # Color based on P&L
                    if i == 17:  # We'll color later based on actual values
                        pass
        
        # Payoff chart
        chart = LineChart()
        chart.title = "P&L Payoff Diagram"
        chart.style = 10
        chart.y_axis.title = 'P&L ($)'
        chart.x_axis.title = 'Copper Price'
        
        data = Reference(ws, min_col=3, min_row=16, max_row=16+len(prices))
        cats = Reference(ws, min_col=1, min_row=17, max_row=16+len(prices))
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        
        ws.add_chart(chart, "G3")
        
        # Column widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 12
        
        print("✓ Created directional trade model")
    
    def create_spread_trade_model(self,
                                 metal1="Copper",
                                 metal2="Aluminum",
                                 entry_ratio=3.76,
                                 target_ratio=4.00,
                                 stop_ratio=3.60,
                                 notional=500000):
        """
        Create spread trade pricing model
        """
        ws = self.wb.create_sheet("Spread Trade")
        
        # Header
        ws['A1'] = 'SPREAD TRADE PRICING MODEL'
        ws['A1'].font = Font(size=16, bold=True, color='FFFFFF')
        ws['A1'].fill = PatternFill(start_color='10b981', end_color='10b981', fill_type='solid')
        ws.merge_cells('A1:F1')
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 30
        
        # Trade parameters
        ws['A3'] = 'SPREAD PARAMETERS'
        ws['A3'].font = Font(bold=True, size=12)
        
        params = [
            ['Long:', metal1],
            ['Short:', metal2],
            ['Entry Ratio:', entry_ratio],
            ['Target Ratio:', target_ratio],
            ['Stop Ratio:', stop_ratio],
            ['Notional (USD):', notional],
            ['', ''],
            ['Upside:', '=(C7-C6)/C6'],
            ['Downside:', '=(C8-C6)/C6']
        ]
        
        for i, (label, value) in enumerate(params, start=4):
            ws[f'A{i}'] = label
            ws[f'A{i}'].font = Font(bold=True)
            ws[f'C{i}'] = value
            if i >= 11:
                ws[f'C{i}'].number_format = '0.00%'
        
        # Spread scenarios
        ws['A15'] = 'SPREAD SCENARIO ANALYSIS'
        ws['A15'].font = Font(bold=True, size=12)
        
        headers = ['Spread Ratio', '% from Entry', 'P&L ($)', 'Return %']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=16, column=col)
            cell.value = header
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='10b981', end_color='10b981', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')
        
        # Spread ratios
        ratios = np.arange(3.4, 4.3, 0.1)
        
        for i, ratio in enumerate(ratios, start=17):
            ws[f'A{i}'] = round(ratio, 2)
            ws[f'B{i}'] = f'=(A{i}-$C$6)/$C$6'
            ws[f'C{i}'] = f'=(A{i}-$C$6)/$C$6*$C$9'
            ws[f'D{i}'] = f'=C{i}/$C$9'
            
            ws[f'B{i}'].number_format = '0.00%'
            ws[f'C{i}'].number_format = '$#,##0'
            ws[f'D{i}'].number_format = '0.00%'
        
        # Chart
        chart = LineChart()
        chart.title = "Spread P&L Profile"
        chart.style = 12
        chart.y_axis.title = 'P&L ($)'
        chart.x_axis.title = f'{metal1}/{metal2} Ratio'
        
        data = Reference(ws, min_col=3, min_row=16, max_row=16+len(ratios))
        cats = Reference(ws, min_col=1, min_row=17, max_row=16+len(ratios))
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        
        ws.add_chart(chart, "G3")
        
        print("✓ Created spread trade model")
    
    def create_option_payoff_model(self,
                                  option_type="Call",
                                  strike=8800,
                                  premium=150,
                                  notional=1000000):
        """
        Create option payoff model (conceptual)
        """
        ws = self.wb.create_sheet("Option Payoff")
        
        # Header
        ws['A1'] = f'{option_type.upper()} OPTION PAYOFF MODEL'
        ws['A1'].font = Font(size=16, bold=True, color='FFFFFF')
        ws['A1'].fill = PatternFill(start_color='eab308', end_color='eab308', fill_type='solid')
        ws.merge_cells('A1:E1')
        ws['A1'].alignment = Alignment(horizontal='center')
        ws.row_dimensions[1].height = 30
        
        # Parameters
        ws['A3'] = 'OPTION PARAMETERS'
        ws['A3'].font = Font(bold=True, size=12)
        
        params = [
            ['Type:', option_type],
            ['Strike Price:', strike],
            ['Premium Paid:', premium],
            ['Notional:', notional],
            ['Break-even:', f'=$C$5+$C$6' if option_type == 'Call' else f'=$C$5-$C$6']
        ]
        
        for i, (label, value) in enumerate(params, start=4):
            ws[f'A{i}'] = label
            ws[f'A{i}'].font = Font(bold=True)
            ws[f'C{i}'] = value
        
        # Payoff table
        ws['A11'] = 'PAYOFF ANALYSIS'
        ws['A11'].font = Font(bold=True, size=12)
        
        headers = ['Spot Price', 'Intrinsic Value', 'Net P&L', 'Return %']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=12, column=col)
            cell.value = header
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='eab308', end_color='eab308', fill_type='solid')
        
        prices = np.arange(strike - 600, strike + 800, 100)
        
        for i, price in enumerate(prices, start=13):
            ws[f'A{i}'] = price
            if option_type == 'Call':
                ws[f'B{i}'] = f'=MAX(A{i}-$C$5,0)'
            else:
                ws[f'B{i}'] = f'=MAX($C$5-A{i},0)'
            ws[f'C{i}'] = f'=B{i}-$C$6'
            ws[f'D{i}'] = f'=C{i}/$C$6'
            
            ws[f'C{i}'].number_format = '$#,##0'
            ws[f'D{i}'].number_format = '0.00%'
        
        print(f"✓ Created {option_type} option model")
    
    def save(self, filename='metals_pricing_models.xlsx'):
        """
        Save Excel workbook
        """
        self.wb.save(filename)
        print(f"\n✓ Saved Excel pricing models: {filename}")


# Example usage
if __name__ == "__main__":
    excel = ExcelPricingModel()
    
    # Create all models
    excel.create_directional_trade_model(
        trade_name="Long Copper",
        entry_price=8650,
        target_price=9200,
        stop_price=8400,
        notional=1000000
    )
    
    excel.create_spread_trade_model(
        metal1="Copper",
        metal2="Aluminum",
        entry_ratio=3.76,
        target_ratio=4.00,
        stop_ratio=3.60,
        notional=500000
    )
    
    excel.create_option_payoff_model(
        option_type="Call",
        strike=8800,
        premium=150,
        notional=1000000
    )
    
    excel.save()
    
    print("\n" + "="*60)
    print("Excel models created with:")
    print("  • Dynamic formulas for P&L calculation")
    print("  • Scenario analysis tables")
    print("  • Professional charts")
    print("  • Risk metrics")
    print("="*60)